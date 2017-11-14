#################################################################
#
#
#    diff_excel.py
#
#    example: diff_excel.py
#
#    auther: Devin Li
#    date:    2016/8/29
#    function: compare excel files
#
#################################################################

import sys
from tkinter import *
import tkinter.filedialog
import tkinter.messagebox
import tkinter.scrolledtext
import subprocess
import xlwt
import xlrd
import openpyxl
import os
import threading

# ===================================================================================

root = tkinter.Tk()
root.title('去除黑名单数据')  # 标题
root.geometry('800x600')  # 窗体大小
root.resizable(False, False)  # 固定窗体
text=tkinter.scrolledtext.ScrolledText(root,font=('微软雅黑',10),fg='blue')

ttag = 1

def showMsg(titleInfo, msgInfo):
    tkinter.messagebox.showwarning(titleInfo, msgInfo)

def showText(textContent, r):
    if(r > -1):
        showInfo = textContent + str(r)
    else:
        showInfo = textContent
    text.insert(END, showInfo + '\n')
    text.see(END)
    #print(textContent, r)

def InputFfile():

    if ttag > 1:
        tkinter.messagebox.showwarning("警告", "正在处理文件，请稍后...")
        return

    global file1
    #label = Label(Tk(), text='选择目标文件')
    #label.pack()
    sInfo = "正在打开目标文件,请稍后..."
    showText(sInfo, -1)
    file1 = tkinter.filedialog.askopenfilenames(filetypes=[("xlsx格式,xls格式", "xlsx xls"), ("All files", "*")],
                                               title='选择目标文件')

    #print(file1)
    if len(file1) == 0:
        tkinter.messagebox.showwarning("提示", "未选中文件，请重新选择")
    else:
        for i in range(0,len(file1)):
            sInfo = "正在检查文件: " + os.path.basename(file1[i]) + "请稍后..."
            showText(sInfo, -1)
            try:
                #print(file1,file2)
                if os.path.splitext(file1[i])[1]==".xlsx":
                    wb1 = openpyxl.load_workbook(file1[i])
                else:
                    wb1 = xlrd.open_workbook(file1[i])
            except:
                msgInfo = os.path.basename(file1[i]) + "文件打开异常，请重新选择"
                tkinter.messagebox.showwarning("提示",msgInfo)
            else:
                sInfo = "文件: " + os.path.basename(file1[i]) + "正常，可处理"
                showText(sInfo, -1)
                if os.path.splitext(file1[i])[1] == ".xlsx":
                    wb1.close()

def InputSfile():

    if ttag > 1:
        tkinter.messagebox.showwarning("警告", "正在处理文件，请稍后...")
        return

    global file2, wb2
    # label = Label(Tk(), text='选择黑名单文件')
    # label.pack()
    sInfo = "正在打开黑名单文件,请稍后..."
    showText(sInfo, -1)
    file2 = tkinter.filedialog.askopenfilename(filetypes=[("xlsx格式,xls格式", "xlsx xls"), ("All files", "*")],
                                               title='选择黑名单文件')
    try:
        sInfo = "正在检查黑名单文件，请稍后..."
        showText(sInfo, -1)
        wb2 = xlrd.open_workbook(file2)
    except:
        tkinter.messagebox.showwarning("提示", "黑名单文件异常，请重新选择")
    else:
        sInfo = "黑名单文件正常，可以开始处理"
        showText(sInfo, -1)

# ===================================================================================

# ===================================================================================
def Diff(i):

    if os.path.splitext(file1[i])[1] == ".xlsx":
        wb1 = openpyxl.load_workbook(file1[i])
        sheet1 = wb1.get_sheet_by_name(wb1.sheetnames[0])
        maxrow = sheet1.max_row
    else:
        wb1 = xlrd.open_workbook(file1[i])
        sheet1 = wb1.sheet_by_index(0)
        maxrow = sheet1.nrows

    try:
        sheet2 = wb2.sheet_by_index(0)
        cols2 = sheet2.col_values(0)
    except:
        tkinter.messagebox.showwarning("提示", "请选择黑名单文件")
    else:
        global ttag
        ttag = ttag + 1

        fileP = os.path.basename(file1[i])
        sInfo = "当前处理文件：" + fileP + "处理条数："
        if os.path.splitext(file1[i])[1] == ".xlsx":
            for r in range(1, maxrow+1):
                #print(sInfo,r)
                showText(sInfo, r)
                sheet1_cell_value = sheet1.cell(row=r, column=1)
                if sheet1_cell_value.value in cols2:
                    sheet1.cell(row=r, column=1).value = None
        else:
            for r in range(0, maxrow):
                #print(fileP, r)
                showText(sInfo, r)
                sheet1_cell_value = sheet1.cell_value(r, 0)
                if sheet1_cell_value in cols2:
                    sheet1_cell_value.write(r, 0, None)

        wb1.save(file1[i])
        wb1.close()
        #print(absPath)
        strInfo = "文件：" + os.path.basename(file1[i]) + "已处理完成！"
        showText(strInfo, -1)
        tkinter.messagebox.showinfo("提示", strInfo)

        ttag = ttag - 1

    # ===================================================================================

def Diffs():

    try:
        len(file1)
    except:
        tkinter.messagebox.showwarning("提示", "请选择目标文件")
    else:
        sInfo = "正在准备开始处理文件"
        showText(sInfo, -1)
        for i in range(0,len(file1)):
            sInfo = "开始处理文件：" + os.path.basename(file1[i])
            showText(sInfo, -1)
            th = threading.Thread(target=Diff,args=(i,))
            th.setDaemon(TRUE)
            th.start()

##########################################################################

tkinter.Button(root, text='请选择目标文件', width=30, height=2, command=InputFfile).pack()
tkinter.Button(root, text='请选择黑名单文件', width=30, height=2, command=InputSfile).pack()
tkinter.Button(root, text='开始处理', width=30, height=2, command=Diffs).pack()
text.pack()

root.mainloop()





