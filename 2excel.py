#################################################################
#
#
#    diff_excel.py
#
#    example: diff_excel.py
#
#    auther: Devin Li
#    date:    2016/8/29
#    function: compare excel files
#
#################################################################

import sys
from tkinter import *
import tkinter.filedialog
import subprocess
#import xlwt
#import xlrd
import openpyxl

# ===================================================================================
def Input():
    global file1, file2
    #label = Label(Tk(), text='选择目标文件')
    #label.pack()
    file1 = tkinter.filedialog.askopenfilename(filetypes=[("xlsx格式,xls格式", "xlsx xls"), ("All files", "*")],
                                               title='选择目标文件')

    #label = Label(Tk(), text='选择黑名单文件')
    #label.pack()
    file2 = tkinter.filedialog.askopenfilename(filetypes=[("xlsx格式,xls格式", "xlsx xls"), ("All files", "*")],
                                               title='选择黑名单文件')


# ===================================================================================

# ===================================================================================
def Diff():

    wb1 = openpyxl.load_workbook(file1)
    wb2 = openpyxl.load_workbook(file2)

    sheet1 = wb1.get_sheet_by_name(wb1.sheetnames[0])
    sheet2 = wb2.get_sheet_by_name(wb2.sheetnames[0])

    wb_result = openpyxl.Workbook()  # 新建一个文件，用来保存结果
    sheet_result = wb_result.create_sheet('result', index=0)

    cols2 = sheet2.columns

    for r in range(1, sheet1.max_row):
        sheet1_cell_value = sheet1.cell(row=r,column=1)
        print("当前处理条数:",r)
        r2 = 0
        for rr in range(1,sheet2.max_row):
            #print(sheet1_cell_value.value,sheet2.cell(row=rr,column=1).value)
            if (sheet1_cell_value.value == sheet2.cell(row=rr,column=1).value):
                break
            else:
                r2 = rr
        #print(c)
        if rr >= sheet2.max_row - 1:
            sheet_result._add_cell(sheet1_cell_value)

    wb_result.save('result.xlsx')

    # ===================================================================================

def main():
    # file1 = sys.argv[1]
    # file2 = sys.argv[2]

    Input()
    Diff()


if __name__ == '__main__':
    # if len(sys.argv)<3:
    #    exit()
    main()

